#!/usr/bin/env python3
"""Parse all baseline xlsx files into a unified baselines.json."""

import json, re, sys, os
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
STRATEGIES_PATH = ROOT / "monitor" / "config" / "strategies.json"
OUTPUT_PATH = ROOT / "monitor" / "baselines.json"


def parse_tv_xlsx(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    result = {"format": "tradingview"}

    # Performance sheet
    perf = {}
    if "Performance" in wb.sheetnames:
        ws = wb["Performance"]
        for row in ws.iter_rows(values_only=True):
            if row[0] and isinstance(row[0], str):
                key = row[0].strip()
                perf[key] = {"all_usd": row[1], "all_pct": row[2]}
    result["net_profit"] = _num(perf.get("Net profit", {}).get("all_usd"))
    result["initial_capital"] = _num(perf.get("Initial capital", {}).get("all_usd"))
    result["cagr_pct"] = _num(perf.get("Annualized return (CAGR)", {}).get("all_pct"))
    result["max_dd_c2c_usd"] = _num(perf.get("Max equity drawdown (close-to-close)", {}).get("all_usd"))
    result["max_dd_c2c_pct"] = _num(perf.get("Max equity drawdown (close-to-close)", {}).get("all_pct"))
    result["max_dd_intrabar_pct"] = _num(
        perf.get("Max equity drawdown as % of initial capital (intrabar)", {}).get("all_pct")
    )

    # Trades analysis
    trades = {}
    if "Trades analysis" in wb.sheetnames:
        ws = wb["Trades analysis"]
        for row in ws.iter_rows(values_only=True):
            if row[0] and isinstance(row[0], str):
                trades[row[0].strip()] = {"all_usd": row[1], "all_pct": row[2]}
    result["total_trades"] = _int(trades.get("Total trades", {}).get("all_usd"))
    result["winning_trades"] = _int(trades.get("Winning trades", {}).get("all_usd"))
    result["losing_trades"] = _int(trades.get("Losing trades", {}).get("all_usd"))
    result["win_pct"] = _num(trades.get("Percent profitable", {}).get("all_pct"))

    # Risk-adjusted performance
    risk = {}
    if "Risk-adjusted performance" in wb.sheetnames:
        ws = wb["Risk-adjusted performance"]
        for row in ws.iter_rows(values_only=True):
            if row[0] and isinstance(row[0], str):
                risk[row[0].strip()] = {"all_usd": row[1], "all_pct": row[2]}
    result["sharpe"] = _num(risk.get("Sharpe ratio", {}).get("all_usd"))
    result["sortino"] = _num(risk.get("Sortino ratio", {}).get("all_usd"))
    result["profit_factor"] = _num(risk.get("Profit factor", {}).get("all_usd"))

    # Properties
    props = {}
    if "Properties" in wb.sheetnames:
        ws = wb["Properties"]
        for row in ws.iter_rows(values_only=True):
            if row[0] and row[1] is not None:
                props[str(row[0]).strip()] = str(row[1]).strip()
    result["properties"] = props
    bt_range = props.get("Backtesting range", props.get("Trading range", ""))
    result["backtest_range"] = _parse_tv_range(bt_range)
    params = {k: v for k, v in props.items() if k not in (
        "Trading range", "Backtesting range", "Symbol", "Timeframe",
        "Point value", "Chart type", "Currency", "Tick size", "Precision",
        "Initial capital", "Order size", "Pyramiding", "Commission",
        "Slippage", "Verify price for limit orders",
        "Margin for long positions", "Margin for short positions",
    )}
    result["params"] = params

    # Max consecutive losses — derived from List of trades
    result["max_consec_loss"] = _derive_max_consec_loss_tv(wb)

    # Compute baseline avg daily return for the soft breach rule
    if result["net_profit"] and result["backtest_range"]:
        start = result["backtest_range"].get("start")
        end = result["backtest_range"].get("end")
        if start and end:
            from datetime import datetime
            try:
                d0 = datetime.strptime(start[:10], "%Y-%m-%d")
                d1 = datetime.strptime(end[:10], "%Y-%m-%d")
                days = max((d1 - d0).days, 1)
                trading_days = int(days * 252 / 365)
                result["avg_daily_return"] = result["net_profit"] / max(trading_days, 1)
            except Exception:
                result["avg_daily_return"] = None

    return result


def _derive_max_consec_loss_tv(wb):
    if "List of trades" not in wb.sheetnames:
        return None
    ws = wb["List of trades"]
    header = None
    pnl_col = None
    trade_num_col = 0
    seen_trades = {}
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c).strip() if c else "" for c in row]
            for i, h in enumerate(header):
                if "net p&l" in h.lower() or "p&l" in h.lower():
                    pnl_col = i
                    break
            if pnl_col is None:
                for i, h in enumerate(header):
                    if "net" in h.lower() and "usd" in h.lower():
                        pnl_col = i
                        break
            continue
        if pnl_col is None:
            continue
        trade_num = row[trade_num_col]
        pnl = row[pnl_col]
        if trade_num is not None and pnl is not None:
            seen_trades[trade_num] = _num(pnl)

    pnls = list(seen_trades.values())
    return _max_consec(pnls, lambda x: x is not None and x < 0)


def _max_consec(items, pred):
    max_run = 0
    cur = 0
    for item in items:
        if pred(item):
            cur += 1
            max_run = max(max_run, cur)
        else:
            cur = 0
    return max_run


def parse_mt5_xlsx(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    result = {"format": "mt5"}
    kv = {}
    inputs = {}
    in_inputs = False
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        joined = " ".join(cells)

        if "Inputs:" in joined:
            in_inputs = True
            continue
        if in_inputs:
            for c in cells:
                m = re.match(r"(Inp\w+)\s*=\s*(.+)", c)
                if m:
                    inputs[m.group(1)] = m.group(2)
                elif c.startswith("input group"):
                    continue
                elif c and not c.startswith("Inp"):
                    in_inputs = False

        for ci in range(0, min(len(cells), 12), 4):
            label = cells[ci] if ci < len(cells) else ""
            value = cells[ci + 3] if ci + 3 < len(cells) else ""
            if not label:
                continue
            label = label.rstrip(":").strip()
            if label:
                kv[label] = value

    result["net_profit"] = _extract_num(kv.get("Total Net Profit", ""))
    result["gross_profit"] = _extract_num(kv.get("Gross Profit", ""))
    result["gross_loss"] = _extract_num(kv.get("Gross Loss", ""))

    dd_str = kv.get("Balance Drawdown Maximal", "")
    dd_match = re.search(r"([\d\s]+\.?\d*)\s*\(([\d.]+)%\)", dd_str)
    if dd_match:
        result["max_dd_usd"] = _extract_num(dd_match.group(1))
        result["max_dd_pct"] = float(dd_match.group(2))
    else:
        result["max_dd_usd"] = _extract_num(dd_str)
        result["max_dd_pct"] = None

    result["profit_factor"] = _extract_num(kv.get("Profit Factor", ""))
    result["recovery_factor"] = _extract_num(kv.get("Recovery Factor", ""))
    result["sharpe"] = _extract_num(kv.get("Sharpe Ratio", ""))
    result["total_trades"] = _extract_int(kv.get("Total Trades", ""))
    result["expected_payoff"] = _extract_num(kv.get("Expected Payoff", ""))

    win_pct_match = re.search(r"([\d.]+)%", kv.get("Profit Trades (% of total)", ""))
    result["win_pct"] = float(win_pct_match.group(1)) if win_pct_match else None

    consec_str = kv.get("Maximum consecutive losses ($)", "")
    consec_match = re.match(r"(\d+)", consec_str)
    result["max_consec_loss"] = int(consec_match.group(1)) if consec_match else None

    period_str = kv.get("Period", "")
    result["backtest_range"] = _parse_mt5_period(period_str)
    result["params"] = inputs

    if result["net_profit"] and result["backtest_range"]:
        start = result["backtest_range"].get("start")
        end = result["backtest_range"].get("end")
        if start and end:
            from datetime import datetime
            try:
                d0 = datetime.strptime(start, "%Y-%m-%d")
                d1 = datetime.strptime(end, "%Y-%m-%d")
                days = max((d1 - d0).days, 1)
                trading_days = int(days * 252 / 365)
                result["avg_daily_return"] = result["net_profit"] / max(trading_days, 1)
            except Exception:
                result["avg_daily_return"] = None

    return result


def _parse_tv_range(s):
    m = re.search(r"(\w+ \d+, \d{4})[^—–-]+[—–-]\s*(\w+ \d+, \d{4})", s)
    if m:
        from datetime import datetime
        try:
            start = datetime.strptime(m.group(1), "%b %d, %Y").strftime("%Y-%m-%d")
            end = datetime.strptime(m.group(2), "%b %d, %Y").strftime("%Y-%m-%d")
            return {"start": start, "end": end}
        except Exception:
            pass
    return {"start": None, "end": None, "raw": s}


def _parse_mt5_period(s):
    m = re.search(r"(\d{4}\.\d{2}\.\d{2})\s*-\s*(\d{4}\.\d{2}\.\d{2})", s)
    if m:
        start = m.group(1).replace(".", "-")
        end = m.group(2).replace(".", "-")
        return {"start": start, "end": end}
    return {"start": None, "end": None, "raw": s}


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace(" ", "").strip())
    except (ValueError, TypeError):
        return None


def _int(v):
    n = _num(v)
    return int(n) if n is not None else None


def _extract_num(s):
    if not s:
        return None
    s = str(s).replace(" ", "").replace(",", "")
    m = re.search(r"-?[\d.]+", s)
    return float(m.group()) if m else None


def _extract_int(s):
    n = _extract_num(s)
    return int(n) if n is not None else None


def main():
    with open(STRATEGIES_PATH) as f:
        strategies = json.load(f)["strategies"]

    baselines = {}
    for strat in strategies:
        sid = strat["id"]
        xlsx_rel = strat["baseline_xlsx"]
        xlsx_path = ROOT / xlsx_rel
        if not xlsx_path.exists():
            print(f"WARNING: {xlsx_path} not found, skipping {sid}")
            continue

        print(f"Parsing {sid} <- {xlsx_rel}")
        if strat["venue"] == "tradingview":
            bl = parse_tv_xlsx(xlsx_path)
        else:
            bl = parse_mt5_xlsx(xlsx_path)

        bl["strategy_id"] = sid
        bl["venue"] = strat["venue"]
        bl["symbol"] = strat["symbol"]
        bl["source_file"] = strat["source_file"]
        baselines[sid] = bl

    with open(OUTPUT_PATH, "w") as f:
        json.dump(baselines, f, indent=2, default=str)
    print(f"\nWrote {len(baselines)} baselines to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
